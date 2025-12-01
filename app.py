import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import tempfile
import os
from io import BytesIO

st.title("Sales data merging")

uploaded_file = st.file_uploader("Fichier Excel", type=["xlsx", "xlsm"])

if uploaded_file and st.button("Générer"):
    
    temp_dir = tempfile.mkdtemp()
    input_path = os.path.join(temp_dir, uploaded_file.name)
    
    with open(input_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    
    wb = load_workbook(input_path, data_only=True)
    p_sheets = [name for name in wb.sheetnames if name.startswith("P.")]
    
    all_data = []
    
    for sheet_name in p_sheets:
        sheet = wb[sheet_name]
        
        product_name = sheet.cell(1, 1).value
        if not product_name:
            continue
        
        plant_origins = []
        for col in range(5, 30):
            origin = sheet.cell(2, col).value
            if not origin or str(origin).strip() == "":
                break
            
            priority_val = sheet.cell(3, col).value
            
            plant_origins.append({
                'origin': str(origin).strip(),
                'priority': str(priority_val).strip() if priority_val and str(priority_val).strip() != "None" else "",
                'col': col
            })
        
        for row in range(4, 250):
            country = sheet.cell(row, 3).value
            if not country or str(country).strip() == "":
                break
            
            person_in_charge = sheet.cell(row, 1).value
            area = sheet.cell(row, 2).value
            ranking = sheet.cell(row, 4).value
            
            for plant in plant_origins:
                status_val = sheet.cell(row, plant['col']).value
                
                all_data.append({
                    'Product name': product_name,
                    'Person in charge': str(person_in_charge).strip() if person_in_charge and str(person_in_charge).strip() != "None" else "",
                    ' Sales Area': str(area).strip() if area and str(area).strip() != "None" else "",
                    'Sales country': str(country).strip(),
                    'Sales ranking': str(ranking).strip() if ranking else "",
                    'Plant origin': plant['origin'],
                    'Priority plant origin': plant['priority'],
                    'Registration status': str(status_val).strip() if status_val and str(status_val).strip() != "None" else ""
                })
    
    wb.close()
    
    df = pd.DataFrame(all_data)
    
    st.success(f"✅ {len(df)} lignes")
    st.dataframe(df.head(10))
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='All data HM', index=False)
    output.seek(0)
    
    st.download_button(
        label="📥 Télécharger",
        data=output,
        file_name="All_data_HM.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    os.remove(input_path)
    os.rmdir(temp_dir)